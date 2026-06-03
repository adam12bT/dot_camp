import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from collections import Counter
import sys, pickle
sys.path.insert(0, '/home/claude')
from engine import evaluate, MATURITY_OPTIONS, MATURITY_SCORES

wb = load_workbook('/mnt/user-data/uploads/Selection_des_startups_Dot_Camp_5.xlsx')

MATURITY_MAP = {
    'Idée': 'Idea',
    'POC finalisé': 'POC finalized',
    'MVP fonctionnel': 'Functional MVP',
    'MVP en cours de test': 'MVP currently being tested',
    'Go To Market (Premières ventes)': 'Go To Market (Early sales)',
    'Produit / Service sur le marché': 'Product/Service on the market',
    'Produit / Service en expansion internationale': 'International Expansion',
    'Functional MVP': 'Functional MVP',
    'MVP currently being tested': 'MVP currently being tested',
    'POC finalized': 'POC finalized',
    'Go To Market (Early sales)': 'Go To Market (Early sales)',
    'Product/Service on the market': 'Product/Service on the market',
    'International Expansion': 'International Expansion',
    'Idea': 'Idea',
}

AGE_MAP = {
    'Moins de 2 ans': 'Less than 2 years',
    'Less than  2 years': 'Less than 2 years',
    'Less than 2 years': 'Less than 2 years',
    '2-5 ans': '2-5 years',
    '2\u20135 ans': '2\u20135 years',
    '2\u20135 years': '2\u20135 years',
    '5-7 ans': '5-7 years',
    '5\u20137 ans': '5\u20137 years',
    '5\u20137 years': '5\u20137 years',
    'Plus de 7 ans': 'More than 7 years',
    'More than 7 years': 'More than 7 years',
}

def parse_bool(v):
    if isinstance(v, bool): return v
    if isinstance(v, (int, float)): return bool(v)
    s = str(v).strip().lower() if v else ''
    return s in {'yes', 'oui', 'true', '1', 'vrai', '1.0'}

def parse_range(v):
    s = str(v).strip() if v else 'None'
    low = s.lower()
    if low in {'aucun', 'none', '0', 'nan', '', 'null'}: return 'None'
    for k in ['1\u20132', '3\u20135', '6\u201310', '+10']:
        if k in s: return k
    if '+10' in s or "'+10" in s or "'+10" in s: return '+10'
    return 'None'

def extract_apps(ws):
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
    def col(name):
        for i, h in enumerate(headers):
            if h and name.lower() in str(h).lower():
                return i + 1
        return None

    name_col = col('startup name') or col('nom de votre startup')
    age_col  = col('age of your startup') or col('ge de votre startup')
    legal_col = col('legally established') or col('cr\u00e9\u00e9e juridiquement')
    ft_col   = col('full-time on the startup') or col('temps plein sur la startup')
    res_col  = col('resident of tunisia') or col('r\u00e9sident en tunisie')
    emp_col  = col('total number of employees') or col('nombre total actuel de collaborateurs')
    sal_col  = col('how many employees are there among') or col('nombre de salari\u00e9s')
    mix_col  = col('gender-mixed') or col('votre \u00e9quipe est-elle mixte')
    mat_col  = col('how advanced is your solution') or col('niveau de maturit\u00e9')
    cli_col  = col('number of customers/users') or col('nombre actuel de clients')
    rev_col  = col('generating revenue') or col('g\u00e9n\u00e8re t-elle actuellement des revenus')
    label_col = col('startup act label') or col('label startup act')
    fund_col = col('raised funding') or col('b\u00e9n\u00e9fici\u00e9 d')
    prog_col = col('incubation or acceleration program') or col('programme d\u2019incubation')
    sec_col  = col('sector of your startup') or col("secteur d'activit\u00e9")
    gov_col  = col('governorate where') or col("gouvernorat d")

    apps = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[name_col-1] if name_col else None
        if not name or str(name).startswith('#') or str(name).strip() == '':
            continue
        age_raw = str(row[age_col-1] or '').strip() if age_col else ''
        age = AGE_MAP.get(age_raw, age_raw)
        # Normalize age en dashes
        age = age.replace('2-5', '2\u20135').replace('5-7', '5\u20137')
        
        mat_raw = str(row[mat_col-1] or 'Idea').strip() if mat_col else 'Idea'
        mat = MATURITY_MAP.get(mat_raw, 'Idea')

        apps.append({
            'startup_name': str(name).strip(),
            'age': age,
            'sector': str(row[sec_col-1] or '').strip() if sec_col else '',
            'governorate': str(row[gov_col-1] or '').strip() if gov_col else '',
            'legally_created': parse_bool(row[legal_col-1]) if legal_col else False,
            'full_time_founder': parse_bool(row[ft_col-1]) if ft_col else False,
            'founder_in_tunisia': parse_bool(row[res_col-1]) if res_col else False,
            'total_employees': parse_range(row[emp_col-1]) if emp_col else 'None',
            'salaried_employees': parse_range(row[sal_col-1]) if sal_col else 'None',
            'gender_mixed': parse_bool(row[mix_col-1]) if mix_col else False,
            'maturity': mat,
            'num_clients': parse_range(row[cli_col-1]) if cli_col else 'None',
            'generating_revenue': parse_bool(row[rev_col-1]) if rev_col else False,
            'startup_label': parse_bool(row[label_col-1]) if label_col else False,
            'raised_funding': parse_bool(row[fund_col-1]) if fund_col else False,
            'participated_programs': parse_bool(row[prog_col-1]) if prog_col else False,
        })
    return apps

ws_en = wb['All Anglais']
ws_fr = wb['All Français']

apps_en = extract_apps(ws_en)
apps_fr = extract_apps(ws_fr)
print(f"EN: {len(apps_en)}, FR: {len(apps_fr)}")

all_apps = apps_en + apps_fr
seen = set()
unique_apps = []
for a in all_apps:
    key = a['startup_name'].lower().strip()
    if key not in seen:
        seen.add(key)
        unique_apps.append(a)
print(f"Unique: {len(unique_apps)}")

# Get old names from Final selection
ws_final = wb['Final selection']
existing_names = set()
for row in ws_final.iter_rows(min_row=2, values_only=True):
    if row[0] and str(row[0]).strip():
        existing_names.add(str(row[0]).strip().lower())
print(f"Existing final: {len(existing_names)}")

results = [(a, evaluate(a)) for a in unique_apps]
old_results = [(a, r) for a, r in results if a['startup_name'].lower().strip() in existing_names]
new_results = [(a, r) for a, r in results if a['startup_name'].lower().strip() not in existing_names]
print(f"Old: {len(old_results)}, New: {len(new_results)}")

# ── Styles ──────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="0D1B2A")
HEADER_FONT  = Font(bold=True, color="00E5A0", name="Arial", size=11)
SEP_FILL     = PatternFill("solid", fgColor="1E2D45")
SEP_FONT     = Font(bold=True, color="F5C842", name="Arial", size=10, italic=True)
OLD_FILL     = PatternFill("solid", fgColor="E8F4FD")
NEW_FILL     = PatternFill("solid", fgColor="F0FDF4")
DEFAULT_FONT = Font(name="Arial", size=10)
DEFAULT_ALIGN = Alignment(vertical="center", wrap_text=False)
THIN_BORDER = Border(
    left=Side(style="thin", color="C0C0C0"),
    right=Side(style="thin", color="C0C0C0"),
    top=Side(style="thin", color="C0C0C0"),
    bottom=Side(style="thin", color="C0C0C0"),
)
DECISION_FILLS = {
    "Selected \u2605": PatternFill("solid", fgColor="00E5A0"),
    "Selected":      PatternFill("solid", fgColor="00B87A"),
    "Shortlisted \u2713": PatternFill("solid", fgColor="F5C842"),
    "Shortlisted":   PatternFill("solid", fgColor="E5A800"),
    "Rejected":      PatternFill("solid", fgColor="FF4D6D"),
}
DECISION_FONTS = {
    "Selected \u2605": Font(bold=True, color="001A10", name="Arial"),
    "Selected":      Font(bold=True, color="001A10", name="Arial"),
    "Shortlisted \u2713": Font(bold=True, color="1A1000", name="Arial"),
    "Shortlisted":   Font(bold=True, color="1A1000", name="Arial"),
    "Rejected":      Font(bold=True, color="FFFFFF", name="Arial"),
}
FILTER_PASS = PatternFill("solid", fgColor="E6FFF7")
FILTER_FAIL = PatternFill("solid", fgColor="FFE4EA")

COLS = [
    ("Startup Name", 24),
    ("Sector", 18),
    ("Governorate", 14),
    ("Age", 18),
    ("Final Decision", 18),
    ("Filter 1", 16),
    ("Filter 2", 16),
    ("Filter 3", 18),
    ("Filter 4", 18),
    ("Bonus Score", 13),
    ("Bonus Details", 48),
    ("Rejection Reason", 36),
]

def write_header(ws):
    ws.row_dimensions[1].height = 30
    for ci, (h, w) in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w

def write_sep(ws, row_idx, label):
    ws.row_dimensions[row_idx].height = 22
    for ci in range(1, len(COLS)+1):
        c = ws.cell(row=row_idx, column=ci)
        c.fill = SEP_FILL
        c.font = SEP_FONT
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="left" if ci == 1 else "center", vertical="center")
    ws.cell(row=row_idx, column=1).value = label

def write_row(ws, row_idx, a, r, bg_fill=None):
    ws.row_dimensions[row_idx].height = 18
    dec = r.final_decision
    row_data = [
        r.startup_name,
        a.get('sector', ''),
        a.get('governorate', ''),
        a.get('age', ''),
        dec,
        r.filter1_result,
        r.filter2_result,
        r.filter3_result,
        r.filter4_result,
        r.bonus_score,
        ", ".join(r.bonus_details),
        r.rejection_reason,
    ]
    for ci, val in enumerate(row_data, 1):
        c = ws.cell(row=row_idx, column=ci, value=val)
        c.border = THIN_BORDER
        c.font = DEFAULT_FONT
        c.alignment = DEFAULT_ALIGN
        if bg_fill and ci not in (5,):
            c.fill = bg_fill
        if ci == 5:
            c.fill = DECISION_FILLS.get(dec, PatternFill())
            c.font = DECISION_FONTS.get(dec, DEFAULT_FONT)
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif ci in (6, 7, 8, 9):
            v = str(val)
            c.fill = FILTER_FAIL if ("Rejected" in v or "Fail" in v) else FILTER_PASS
        elif ci == 10:
            c.alignment = Alignment(horizontal="center", vertical="center")

# ── Build new sheet: Engine Results ─────────────────────────────────────────
if "Engine Results" in wb.sheetnames:
    del wb["Engine Results"]
ws_eng = wb.create_sheet("Engine Results")
ws_eng.freeze_panes = "A2"
write_header(ws_eng)

row = 2
# Old entries section
write_sep(ws_eng, row, f"  ── ORIGINAL APPLICATIONS ({len(old_results)} startups) ──")
row += 1
for a, r in sorted(old_results, key=lambda x: (
    {"Selected ★":0,"Selected":1,"Shortlisted ✓":2,"Shortlisted":3,"Rejected":4}.get(x[1].final_decision, 5),
    -x[1].bonus_score
)):
    write_row(ws_eng, row, a, r, OLD_FILL)
    row += 1

# Separator between old and new
row += 1
write_sep(ws_eng, row, f"  ── NEW APPLICATIONS ({len(new_results)} startups) ──")
row += 1
for a, r in sorted(new_results, key=lambda x: (
    {"Selected ★":0,"Selected":1,"Shortlisted ✓":2,"Shortlisted":3,"Rejected":4}.get(x[1].final_decision, 5),
    -x[1].bonus_score
)):
    write_row(ws_eng, row, a, r, NEW_FILL)
    row += 1

ws_eng.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"

# ── Build Dashboard sheet ────────────────────────────────────────────────────
if "Dashboard" in wb.sheetnames:
    del wb["Dashboard"]
ws_dash = wb.create_sheet("Dashboard", 0)

all_results = [r for _, r in results]
old_res_only = [r for _, r in old_results]
new_res_only = [r for _, r in new_results]

DECISIONS = ["Selected \u2605", "Selected", "Shortlisted \u2713", "Shortlisted", "Rejected"]

def dash_header(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN_BORDER
    return c

def dash_val(ws, row, col, val, fill=None, bold=False, fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", size=11, bold=bold)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN_BORDER
    if fill: c.fill = fill
    if fmt: c.number_format = fmt
    return c

# Title
ws_dash.merge_cells("A1:L1")
t = ws_dash.cell(row=1, column=1, value="🚀  Dot Camp 5 – Selection Dashboard")
t.fill = PatternFill("solid", fgColor="0D1B2A")
t.font = Font(bold=True, color="00E5A0", name="Arial", size=16)
t.alignment = Alignment(horizontal="center", vertical="center")
ws_dash.row_dimensions[1].height = 40

# Section: Overview stats
ws_dash.row_dimensions[3].height = 22
ws_dash.merge_cells("A3:L3")
h = ws_dash.cell(row=3, column=1, value="OVERALL STATISTICS")
h.fill = SEP_FILL; h.font = SEP_FONT
h.alignment = Alignment(horizontal="center", vertical="center")

labels = ["Total Applications", "Selected ★", "Selected", "Shortlisted ✓", "Shortlisted", "Rejected"]
totals = [len(all_results)] + [Counter(r.final_decision for r in all_results).get(d, 0) for d in DECISIONS]
fill_colors = ["2196F3", "00E5A0", "00B87A", "F5C842", "E5A800", "FF4D6D"]
text_colors = ["FFFFFF", "001A10", "001A10", "1A1000", "1A1000", "FFFFFF"]

for i, (lbl, val, fc, tc) in enumerate(zip(labels, totals, fill_colors, text_colors)):
    col = i*2 + 1
    ws_dash.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
    ws_dash.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)
    lc = ws_dash.cell(row=4, column=col, value=lbl)
    lc.fill = PatternFill("solid", fgColor=fc)
    lc.font = Font(bold=True, color=tc, name="Arial", size=10)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    lc.border = THIN_BORDER
    vc = ws_dash.cell(row=5, column=col, value=val)
    vc.fill = PatternFill("solid", fgColor="F8F8F8")
    vc.font = Font(bold=True, name="Arial", size=18)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    vc.border = THIN_BORDER
    ws_dash.row_dimensions[4].height = 24
    ws_dash.row_dimensions[5].height = 36

# Old vs New comparison table
ws_dash.row_dimensions[7].height = 22
ws_dash.merge_cells("A7:L7")
h2 = ws_dash.cell(row=7, column=1, value="OLD vs NEW APPLICATIONS BREAKDOWN")
h2.fill = SEP_FILL; h2.font = SEP_FONT
h2.alignment = Alignment(horizontal="center", vertical="center")

headers_cmp = ["Decision", "Original Count", "Original %", "New Count", "New %", "Total"]
col_widths_cmp = [20, 16, 14, 12, 12, 10]
for ci, (h, w) in enumerate(zip(headers_cmp, col_widths_cmp), 1):
    dash_header(ws_dash, 8, ci, h)
    ws_dash.column_dimensions[get_column_letter(ci)].width = w

c_old = Counter(r.final_decision for r in old_res_only)
c_new = Counter(r.final_decision for r in new_res_only)
all_dec_order = ["Selected \u2605", "Selected", "Shortlisted \u2713", "Shortlisted", "Rejected"]
row_d = 9
for dec in all_dec_order:
    o = c_old.get(dec, 0)
    n = c_new.get(dec, 0)
    o_pct = o/len(old_res_only)*100 if old_res_only else 0
    n_pct = n/len(new_res_only)*100 if new_res_only else 0
    f = DECISION_FILLS.get(dec, PatternFill())
    fn = DECISION_FONTS.get(dec, DEFAULT_FONT)
    c = ws_dash.cell(row=row_d, column=1, value=dec)
    c.fill = f; c.font = fn; c.alignment = Alignment(horizontal="center", vertical="center"); c.border = THIN_BORDER
    for ci2, val in enumerate([o, f"{o_pct:.1f}%", n, f"{n_pct:.1f}%", o+n], 2):
        cc = ws_dash.cell(row=row_d, column=ci2, value=val)
        cc.font = Font(name="Arial", size=10)
        cc.alignment = Alignment(horizontal="center", vertical="center")
        cc.border = THIN_BORDER
    ws_dash.row_dimensions[row_d].height = 20
    row_d += 1

# Sector breakdown
ws_dash.row_dimensions[row_d+1].height = 22
ws_dash.merge_cells(start_row=row_d+1, start_column=1, end_row=row_d+1, end_column=6)
hs = ws_dash.cell(row=row_d+1, column=1, value="TOP SECTORS")
hs.fill = SEP_FILL; hs.font = SEP_FONT
hs.alignment = Alignment(horizontal="center", vertical="center")

dash_header(ws_dash, row_d+2, 1, "Sector")
dash_header(ws_dash, row_d+2, 2, "Total")
dash_header(ws_dash, row_d+2, 3, "Selected")
dash_header(ws_dash, row_d+2, 4, "Shortlisted")
dash_header(ws_dash, row_d+2, 5, "Rejected")
dash_header(ws_dash, row_d+2, 6, "% Selected")
ws_dash.column_dimensions['A'].width = 22

sector_data = {}
for a, r in results:
    sec = a.get('sector') or 'Other'
    if sec not in sector_data:
        sector_data[sec] = {'total':0,'selected':0,'shortlisted':0,'rejected':0}
    sector_data[sec]['total'] += 1
    if 'Selected' in r.final_decision:
        sector_data[sec]['selected'] += 1
    elif 'Shortlisted' in r.final_decision:
        sector_data[sec]['shortlisted'] += 1
    else:
        sector_data[sec]['rejected'] += 1

sector_sorted = sorted(sector_data.items(), key=lambda x: -x[1]['total'])[:12]
sr = row_d + 3
for sec, d in sector_sorted:
    ws_dash.cell(row=sr, column=1, value=sec).border = THIN_BORDER
    ws_dash.cell(row=sr, column=1).font = Font(name="Arial", size=10)
    ws_dash.cell(row=sr, column=2, value=d['total']).border = THIN_BORDER
    ws_dash.cell(row=sr, column=2).alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.cell(row=sr, column=3, value=d['selected']).border = THIN_BORDER
    ws_dash.cell(row=sr, column=3).fill = PatternFill("solid", fgColor="E6FFF7")
    ws_dash.cell(row=sr, column=3).alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.cell(row=sr, column=4, value=d['shortlisted']).border = THIN_BORDER
    ws_dash.cell(row=sr, column=4).fill = PatternFill("solid", fgColor="FFFDE7")
    ws_dash.cell(row=sr, column=4).alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.cell(row=sr, column=5, value=d['rejected']).border = THIN_BORDER
    ws_dash.cell(row=sr, column=5).fill = PatternFill("solid", fgColor="FFEAEE")
    ws_dash.cell(row=sr, column=5).alignment = Alignment(horizontal="center", vertical="center")
    pct = d['selected']/d['total']*100 if d['total'] else 0
    ws_dash.cell(row=sr, column=6, value=f"{pct:.0f}%").border = THIN_BORDER
    ws_dash.cell(row=sr, column=6).alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[sr].height = 18
    sr += 1

# Maturity breakdown
ws_dash.merge_cells(start_row=row_d+1, start_column=8, end_row=row_d+1, end_column=12)
hm = ws_dash.cell(row=row_d+1, column=8, value="MATURITY LEVEL BREAKDOWN")
hm.fill = SEP_FILL; hm.font = SEP_FONT
hm.alignment = Alignment(horizontal="center", vertical="center")

for ci2, h in enumerate(["Maturity", "Count", "Selected", "Shortlisted", "Rejected"], 8):
    dash_header(ws_dash, row_d+2, ci2, h)
ws_dash.column_dimensions['H'].width = 26

mat_data = {}
for a, r in results:
    m = a.get('maturity', 'Unknown')
    if m not in mat_data:
        mat_data[m] = {'total':0,'selected':0,'shortlisted':0,'rejected':0}
    mat_data[m]['total'] += 1
    if 'Selected' in r.final_decision:
        mat_data[m]['selected'] += 1
    elif 'Shortlisted' in r.final_decision:
        mat_data[m]['shortlisted'] += 1
    else:
        mat_data[m]['rejected'] += 1

mat_order = ['Idea','POC finalized','Functional MVP','MVP currently being tested',
             'Go To Market (Early sales)','Product/Service on the market','International Expansion']
mr = row_d + 3
for m in mat_order:
    if m not in mat_data: continue
    d = mat_data[m]
    ws_dash.cell(row=mr, column=8, value=m).border = THIN_BORDER
    ws_dash.cell(row=mr, column=8).font = Font(name="Arial", size=10)
    for ci2, val in enumerate([d['total'], d['selected'], d['shortlisted'], d['rejected']], 9):
        c = ws_dash.cell(row=mr, column=ci2, value=val)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
        fills2 = {9: None, 10: "E6FFF7", 11: "FFFDE7", 12: "FFEAEE"}
        if fills2.get(ci2):
            c.fill = PatternFill("solid", fgColor=fills2[ci2])
    ws_dash.row_dimensions[mr].height = 18
    mr += 1

# Governorate breakdown
gov_start_row = max(sr, mr) + 2
ws_dash.merge_cells(start_row=gov_start_row, start_column=1, end_row=gov_start_row, end_column=6)
hg = ws_dash.cell(row=gov_start_row, column=1, value="APPLICATIONS BY GOVERNORATE (TOP 10)")
hg.fill = SEP_FILL; hg.font = SEP_FONT
hg.alignment = Alignment(horizontal="center", vertical="center")
ws_dash.row_dimensions[gov_start_row].height = 22

for ci2, h in enumerate(["Governorate", "Total", "Selected", "Shortlisted", "Rejected", "Select Rate"], 1):
    dash_header(ws_dash, gov_start_row+1, ci2, h)

gov_data = {}
for a, r in results:
    g = a.get('governorate') or 'Unknown'
    if g not in gov_data:
        gov_data[g] = {'total':0,'selected':0,'shortlisted':0,'rejected':0}
    gov_data[g]['total'] += 1
    if 'Selected' in r.final_decision: gov_data[g]['selected'] += 1
    elif 'Shortlisted' in r.final_decision: gov_data[g]['shortlisted'] += 1
    else: gov_data[g]['rejected'] += 1

gov_sorted = sorted(gov_data.items(), key=lambda x: -x[1]['total'])[:10]
gr = gov_start_row + 2
for gov, d in gov_sorted:
    ws_dash.cell(row=gr, column=1, value=gov).border = THIN_BORDER
    ws_dash.cell(row=gr, column=1).font = Font(name="Arial", size=10)
    for ci2, val in enumerate([d['total'], d['selected'], d['shortlisted'], d['rejected']], 2):
        c = ws_dash.cell(row=gr, column=ci2, value=val)
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center", vertical="center")
    pct = (d['selected'])/d['total']*100 if d['total'] else 0
    c6 = ws_dash.cell(row=gr, column=6, value=f"{pct:.0f}%")
    c6.border = THIN_BORDER; c6.alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[gr].height = 18
    gr += 1

# Rejection reasons analysis
rej_start = max(sr, mr) + 2
ws_dash.merge_cells(start_row=rej_start, start_column=8, end_row=rej_start, end_column=12)
hr = ws_dash.cell(row=rej_start, column=8, value="REJECTION REASONS")
hr.fill = SEP_FILL; hr.font = SEP_FONT
hr.alignment = Alignment(horizontal="center", vertical="center")
ws_dash.row_dimensions[rej_start].height = 22

dash_header(ws_dash, rej_start+1, 8, "Reason")
dash_header(ws_dash, rej_start+1, 9, "Count")
ws_dash.column_dimensions['I'].width = 10

rej_reasons = Counter(r.rejection_reason for _, r in results if r.rejection_reason)
rr = rej_start + 2
for reason, cnt in rej_reasons.most_common(8):
    ws_dash.cell(row=rr, column=8, value=reason).border = THIN_BORDER
    ws_dash.cell(row=rr, column=8).font = Font(name="Arial", size=9)
    ws_dash.cell(row=rr, column=8).fill = PatternFill("solid", fgColor="FFEAEE")
    c9 = ws_dash.cell(row=rr, column=9, value=cnt)
    c9.border = THIN_BORDER; c9.alignment = Alignment(horizontal="center", vertical="center")
    c9.font = Font(bold=True, name="Arial", size=10)
    ws_dash.row_dimensions[rr].height = 18
    ws_dash.column_dimensions['H'].width = 38
    rr += 1

# Bonus score distribution
bonus_start = gr + 2
ws_dash.merge_cells(start_row=bonus_start, start_column=1, end_row=bonus_start, end_column=6)
hb = ws_dash.cell(row=bonus_start, column=1, value="BONUS SCORE DISTRIBUTION (non-rejected)")
hb.fill = SEP_FILL; hb.font = SEP_FONT
hb.alignment = Alignment(horizontal="center", vertical="center")
ws_dash.row_dimensions[bonus_start].height = 22

dash_header(ws_dash, bonus_start+1, 1, "Score Range")
dash_header(ws_dash, bonus_start+1, 2, "Count")
dash_header(ws_dash, bonus_start+1, 3, "Decisions")

non_rej = [(a, r) for a, r in results if not r.is_rejected]
score_buckets = {"0-2": 0, "3-4": 0, "5-6": 0, "7-8": 0, "9+": 0}
for a, r in non_rej:
    s = r.bonus_score
    if s <= 2: score_buckets["0-2"] += 1
    elif s <= 4: score_buckets["3-4"] += 1
    elif s <= 6: score_buckets["5-6"] += 1
    elif s <= 8: score_buckets["7-8"] += 1
    else: score_buckets["9+"] += 1

brow = bonus_start + 2
for rng, cnt in score_buckets.items():
    ws_dash.cell(row=brow, column=1, value=rng).border = THIN_BORDER
    ws_dash.cell(row=brow, column=1).alignment = Alignment(horizontal="center", vertical="center")
    c2 = ws_dash.cell(row=brow, column=2, value=cnt)
    c2.border = THIN_BORDER; c2.alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[brow].height = 18
    brow += 1

# Selected startups list (quick ref)
sel_start = rr + 2
ws_dash.merge_cells(start_row=sel_start, start_column=8, end_row=sel_start, end_column=12)
hs2 = ws_dash.cell(row=sel_start, column=8, value="TOP SELECTED STARTUPS")
hs2.fill = HEADER_FILL
hs2.font = Font(bold=True, color="00E5A0", name="Arial", size=11)
hs2.alignment = Alignment(horizontal="center", vertical="center")
ws_dash.row_dimensions[sel_start].height = 22

dash_header(ws_dash, sel_start+1, 8, "Startup")
dash_header(ws_dash, sel_start+1, 9, "Decision")
dash_header(ws_dash, sel_start+1, 10, "Bonus")
dash_header(ws_dash, sel_start+1, 11, "Sector")
ws_dash.column_dimensions['J'].width = 18
ws_dash.column_dimensions['K'].width = 12
ws_dash.column_dimensions['L'].width = 18

top_selected = sorted(
    [(a, r) for a, r in results if r.final_decision in ("Selected \u2605", "Selected")],
    key=lambda x: (-x[1].bonus_score, x[1].final_decision)
)[:15]

trow = sel_start + 2
for a, r in top_selected:
    c8 = ws_dash.cell(row=trow, column=8, value=r.startup_name)
    c8.border = THIN_BORDER; c8.font = Font(name="Arial", size=10)
    c9 = ws_dash.cell(row=trow, column=9, value=r.final_decision)
    c9.border = THIN_BORDER
    c9.fill = DECISION_FILLS.get(r.final_decision, PatternFill())
    c9.font = DECISION_FONTS.get(r.final_decision, DEFAULT_FONT)
    c9.alignment = Alignment(horizontal="center", vertical="center")
    c10 = ws_dash.cell(row=trow, column=10, value=r.bonus_score)
    c10.border = THIN_BORDER; c10.alignment = Alignment(horizontal="center", vertical="center")
    c11 = ws_dash.cell(row=trow, column=11, value=a.get('sector',''))
    c11.border = THIN_BORDER; c11.font = Font(name="Arial", size=10)
    ws_dash.row_dimensions[trow].height = 18
    trow += 1

wb.save('/mnt/user-data/outputs/Selection_des_startups_Dot_Camp_5_Enhanced.xlsx')
print("DONE - saved to outputs")
print(f"Total: {len(results)}, Old: {len(old_results)}, New: {len(new_results)}")