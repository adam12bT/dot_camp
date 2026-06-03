"""
Dot Camp 5 – Startup Selection System
Streamlit App

Run:  streamlit run app.py
"""

import streamlit as st
import pandas as pd
import io
from copy import copy

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from engine import evaluate, MATURITY_OPTIONS, AGE_OPTIONS, EMPLOYEE_RANGES, CLIENT_RANGES

# ─── Hardcoded Excel path ──────────────────────────────────────────────────────
EXCEL_PATH = "data/Selection des startups Dot Camp 5 (2).xlsx"
# Examples:
#   Windows: "C:/Users/YourName/Documents/Selection_des_startups_Dot_Camp_5.xlsx"
#   Mac/Linux: "/home/yourname/Documents/Selection_des_startups_Dot_Camp_5.xlsx"

# ─── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Dot Camp 5 – Selection Engine",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=DM+Serif+Display&display=swap');
  html, body, [class*="css"] { font-family: 'DM Mono', monospace; }
  h1, h2, h3 { font-family: 'DM Serif Display', serif !important; }
  .stMetric label { font-size: 11px; letter-spacing: 2px; text-transform: uppercase; }
  .decision-badge { display:inline-block; padding:4px 12px; border-radius:5px; font-weight:700; font-size:13px; }
  .badge-selected-star  { background:#00e5a0; color:#001a10; }
  .badge-selected       { background:#00b87a; color:#001a10; }
  .badge-shortlisted-ok { background:#f5c842; color:#1a1000; }
  .badge-shortlisted    { background:#e5a800; color:#1a1000; }
  .badge-rejected       { background:#ff4d6d; color:#fff;    }
</style>
""", unsafe_allow_html=True)

DECISION_COLORS = {
    "Selected ★":    "#00e5a0",
    "Selected":      "#00b87a",
    "Shortlisted ✓": "#f5c842",
    "Shortlisted":   "#e5a800",
    "Rejected":      "#ff4d6d",
}

def color_decision(val):
    c = DECISION_COLORS.get(val, "")
    return f"background-color: {c}22; color: {c}; font-weight: bold;" if c else ""

# ─── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## Dot Camp 5 Selection Engine")
    st.divider()
    st.markdown("### Filter Logic")
    st.markdown("""
| Filter | Rule |
|--------|------|
| F1 | Age: >7y → Rejected |
| F2 | 2-5y must be legally created |
| F3 | Founder in TN + full-time |
| F4 | Maturity ≥ MVP → Selected |
""")

# ─── Main ──────────────────────────────────────────────────────────────────────
st.markdown("#  Dot Camp 5 – Startup Selection Engine")
st.divider()

# ══════════════════════════════════════════════════════════════════════════════
# COLUMN MAPS — exact columns matching the real Excel
# ══════════════════════════════════════════════════════════════════════════════

EN_COLS = {
    "startup_name":          3,
    "age":                   4,
    "governorate":           5,
    "sector":                6,
    "legally_created":       8,
    "founded_year":          9,
    "full_time_founder":     39,
    "founder_in_tunisia":    40,
    "total_employees":       41,
    "salaried_employees":    42,
    "gender_mixed":          43,
    "maturity":              49,
    "num_clients":           50,
    "generating_revenue":    51,
    "hiring_plans":          56,
    "raised_funding":        57,
}

FR_COLS = {
    "startup_name":          3,   # col 1 = form submission ID, col 3 = actual startup name
    "age":                   4,
    "governorate":           5,
    "sector":                6,
    "legally_created":       8,
    "founded_year":          9,
    "full_time_founder":     39,
    "founder_in_tunisia":    40,
    "total_employees":       41,
    "salaried_employees":    42,
    "gender_mixed":          43,
    "maturity":              49,
    "num_clients":           50,
    "generating_revenue":    51,
    "hiring_plans":          56,
    "raised_funding":        57,
}

FS_COLS = {
    "startup_name":       1,
    "age":                2,
    "governorate":        3,
    "sector":             4,
    "legally_created":    6,
    "founded_year":       7,
    "full_time_founder":  8,
    "founder_in_tunisia": 9,
    "total_employees":    10,
    "salaried_employees": 11,
    "gender_mixed":       12,
    "maturity":           19,
    "generating_revenue": 20,
    "raised_funding":     22,
}

# ── Normalisation helpers ──────────────────────────────────────────────────────
MATURITY_MAP = {
    'Idée': 'Idea', 'POC finalisé': 'POC finalized',
    'MVP fonctionnel': 'Functional MVP',
    'MVP en cours de test': 'MVP currently being tested',
    'Go To Market (Premières ventes)': 'Go To Market (Early sales)',
    'Produit / Service sur le marché': 'Product/Service on the market',
    'Produit / Service en expansion internationale': 'International Expansion',
}
for v in list(MATURITY_OPTIONS): MATURITY_MAP[v] = v

AGE_MAP = {
    'Moins de 2 ans': 'Less than 2 years',
    'Less than  2 years': 'Less than 2 years',
    '2-5 ans': '2–5 years', '2–5 ans': '2–5 years',
    '5-7 ans': '5–7 years', '5–7 ans': '5–7 years',
    'Plus de 7 ans': 'More than 7 years',
}
for v in AGE_OPTIONS: AGE_MAP[v] = v

def _parse_bool(v):
    if isinstance(v, bool): return v
    if isinstance(v, (int, float)): return bool(v)
    s = str(v).strip().lower() if v else ''
    return s in {'yes', 'oui', 'true', '1', 'vrai', '1.0'}

def _parse_range(v):
    s = str(v).strip() if v else 'None'
    if s.lower() in {'aucun', 'none', '0', 'nan', '', 'null'}: return 'None'
    for k in ['1–2', '3–5', '6–10', '+10']:
        if k in s or k.replace('–', '-') in s: return k
    return 'None'

def _bool_str(v): return "Yes" if v else "No"

def _copy_style(src, dst):
    if src.has_style:
        dst.font       = copy(src.font)
        dst.border     = copy(src.border)
        dst.fill       = copy(src.fill)
        dst.number_format = copy(src.number_format)
        dst.protection = copy(src.protection)
        dst.alignment  = copy(src.alignment)

def _extract_apps(ws, col_map):
    name_col = col_map["startup_name"]
    apps = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[name_col - 1] if len(row) >= name_col else None
        if not name or str(name).strip() == '' or str(name).startswith('#'):
            continue
        age_raw = str(row[col_map["age"] - 1] or '').strip()
        age = AGE_MAP.get(age_raw, age_raw)
        mat_raw = str(row[col_map["maturity"] - 1] or 'Idea').strip()
        mat = MATURITY_MAP.get(mat_raw, 'Idea')
        apps.append({
            'startup_name':          str(name).strip(),
            'age':                   age,
            'sector':                str(row[col_map["sector"] - 1] or '').strip(),
            'governorate':           str(row[col_map["governorate"] - 1] or '').strip(),
            'legally_created':       _parse_bool(row[col_map["legally_created"] - 1]),
            'full_time_founder':     _parse_bool(row[col_map["full_time_founder"] - 1]),
            'founder_in_tunisia':    _parse_bool(row[col_map["founder_in_tunisia"] - 1]),
            'total_employees':       _parse_range(row[col_map["total_employees"] - 1]),
            'salaried_employees':    _parse_range(row[col_map["salaried_employees"] - 1]),
            'gender_mixed':          _parse_bool(row[col_map["gender_mixed"] - 1]),
            'maturity':              mat,
            'num_clients':           _parse_range(row[col_map["num_clients"] - 1]) if "num_clients" in col_map else 'None',
            'generating_revenue':    _parse_bool(row[col_map["generating_revenue"] - 1]),
            'raised_funding':        _parse_bool(row[col_map["raised_funding"] - 1]),
            'participated_programs': False,
            'startup_label':         False,
        })
    return apps

def _last_data_row(ws):
    for r in range(ws.max_row, 1, -1):
        if any(ws.cell(row=r, column=c).value is not None for c in range(1, ws.max_column + 1)):
            return r
    return 1

def _append_to_main_sheet(ws, col_map, app, founded_year=""):
    last_row = _last_data_row(ws)
    new_row  = last_row + 1
    total_cols = ws.max_column
    for c in range(1, total_cols + 1):
        _copy_style(ws.cell(row=last_row, column=c),
                    ws.cell(row=new_row,  column=c))
    def w(field, value):
        if field in col_map:
            ws.cell(row=new_row, column=col_map[field]).value = value
    w("startup_name",       app['startup_name'])
    w("age",                app['age'])
    w("governorate",        app['governorate'])
    w("sector",             app['sector'])
    w("legally_created",    _bool_str(app['legally_created']))
    w("founded_year",       founded_year)
    w("full_time_founder",  _bool_str(app['full_time_founder']))
    w("founder_in_tunisia", _bool_str(app['founder_in_tunisia']))
    w("total_employees",    app['total_employees'])
    w("salaried_employees", app['salaried_employees'])
    w("gender_mixed",       _bool_str(app['gender_mixed']))
    w("maturity",           app['maturity'])
    w("num_clients",        app['num_clients'])
    w("generating_revenue", _bool_str(app['generating_revenue']))
    w("raised_funding",     _bool_str(app['raised_funding']))

def _append_to_final_selection(ws, app, founded_year=""):
    last_row = _last_data_row(ws)
    new_row  = last_row + 1
    total_cols = ws.max_column
    for c in range(1, total_cols + 1):
        _copy_style(ws.cell(row=last_row, column=c),
                    ws.cell(row=new_row,  column=c))
    def w(col, value):
        ws.cell(row=new_row, column=col).value = value
    w(FS_COLS["startup_name"],       app['startup_name'])
    w(FS_COLS["age"],                app['age'])
    w(FS_COLS["governorate"],        app['governorate'])
    w(FS_COLS["sector"],             app['sector'])
    w(FS_COLS["legally_created"],    _bool_str(app['legally_created']))
    w(FS_COLS["founded_year"],       founded_year)
    w(FS_COLS["full_time_founder"],  _bool_str(app['full_time_founder']))
    w(FS_COLS["founder_in_tunisia"], _bool_str(app['founder_in_tunisia']))
    w(FS_COLS["total_employees"],    app['total_employees'])
    w(FS_COLS["salaried_employees"], app['salaried_employees'])
    w(FS_COLS["gender_mixed"],       _bool_str(app['gender_mixed']))
    w(FS_COLS["maturity"],           app['maturity'])
    w(FS_COLS["generating_revenue"], _bool_str(app['generating_revenue']))
    w(FS_COLS["raised_funding"],     _bool_str(app['raised_funding']))


# ══════════════════════════════════════════════════════════════════════════════
# LOAD FILE
# ══════════════════════════════════════════════════════════════════════════════

try:
    wb = load_workbook(EXCEL_PATH)
except FileNotFoundError:
    st.error(f"❌ File not found: `{EXCEL_PATH}`\n\nPlease update the `EXCEL_PATH` variable at the top of `app.py`.")
    st.stop()
except Exception as e:
    st.error(f"❌ Failed to load Excel file: {e}")
    st.stop()

uploaded_xl_name = EXCEL_PATH.split("/")[-1].split("\\")[-1]  # works for both Windows and Unix paths

has_en = "All Anglais" in wb.sheetnames
has_fr = "All Français" in wb.sheetnames
has_fs = "Final selection" in wb.sheetnames

if not has_en and not has_fr:
    st.error("❌ Could not find 'All Anglais' or 'All Français' sheet.")
    st.stop()

# Extract existing apps for duplicate check
existing_names = set()
if has_en:
    for a in _extract_apps(wb["All Anglais"], EN_COLS):
        existing_names.add(a['startup_name'].lower().strip())
if has_fr:
    for a in _extract_apps(wb["All Français"], FR_COLS):
        existing_names.add(a['startup_name'].lower().strip())

st.success(f"✅ Loaded **{len(existing_names)} existing startups** from `{uploaded_xl_name}`")

# ── Show existing startups ────────────────────────────────────────────────────
with st.expander("👁 View existing startups", expanded=False):
    all_apps = []
    if has_en:
        all_apps += _extract_apps(wb["All Anglais"], EN_COLS)
    if has_fr:
        all_apps += _extract_apps(wb["All Français"], FR_COLS)
    if all_apps:
        results_existing = [evaluate(a) for a in all_apps]
        df_ex = pd.DataFrame([{
            "Startup": r.startup_name,
            "Decision": r.final_decision,
            "Bonus": r.bonus_score,
            "Age": a['age'],
            "Maturity": a['maturity'],
        } for a, r in zip(all_apps, results_existing)])
        st.dataframe(df_ex.style.applymap(color_decision, subset=["Decision"]),
                     use_container_width=True, height=400)

# ── Add new startup ───────────────────────────────────────────────────────────
st.divider()
st.markdown("### ➕ Add new startup(s)")

tab_manual, tab_csv, tab_json, tab_override = st.tabs(["✏️ Manual Entry", "📄 Google Form CSV", "📋 Typeform JSON", "🎛️ Manual Override"])

def _process_batch(new_apps, target_sheet, col_map, wb, existing_names, has_fs, xl_name):
    unique = [(a, evaluate(a)) for a in new_apps
              if a["startup_name"].lower().strip() not in existing_names]
    if not unique:
        st.warning("⚠️ All applications already exist in the file — nothing added.")
        return
    st.markdown(f"#### 🔍 Results for {len(unique)} new application(s)")
    df = pd.DataFrame([{
        "Startup":     r.startup_name,
        "Decision":    r.final_decision,
        "Bonus":       r.bonus_score,
        "F1 Age":      r.filter1_result,
        "F2 Legal":    r.filter2_result,
        "F3 Founder":  r.filter3_result,
        "F4 Maturity": r.filter4_result,
    } for _, r in unique])
    st.dataframe(df.style.applymap(color_decision, subset=["Decision"]),
                 use_container_width=True, height=300)
    selected_names = []
    for app, result in unique:
        _append_to_main_sheet(wb[target_sheet], col_map, app)
        if result.final_decision in ("Selected ★", "Selected") and has_fs:
            _append_to_final_selection(wb["Final selection"], app)
            selected_names.append(app["startup_name"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    msg = f"✅ **{len(unique)}** startup(s) added to `{target_sheet}`"
    if selected_names:
        msg += f" · **{len(selected_names)}** also added to `Final selection`: " + ", ".join(selected_names)
    st.success(msg)
    stem = xl_name.replace(".xlsx", "")
    st.download_button(
        "⬇️ Download Updated Excel",
        data=buf.getvalue(),
        file_name=f"{stem}_Updated.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

with tab_csv:
    from mock_data import gform_csv_to_apps
    st.info("Upload a CSV exported from **Google Form**.")
    csv_lang = st.radio("Target sheet", ["English (All Anglais)", "Français (All Français)"],
                        horizontal=True, key="csv_lang")
    new_csv = st.file_uploader("Upload CSV", type=["csv"], key="new_csv")
    if new_csv:
        new_apps_csv = gform_csv_to_apps(new_csv.read().decode("utf-8"))
        st.info(f"Found **{len(new_apps_csv)}** application(s) in CSV")
        if st.button("✅ Evaluate & Add all", key="btn_csv", type="primary"):
            _process_batch(
                new_apps_csv,
                "All Anglais" if csv_lang.startswith("English") else "All Français",
                EN_COLS if csv_lang.startswith("English") else FR_COLS,
                wb, existing_names, has_fs, uploaded_xl_name,
            )

with tab_json:
    from mock_data import typeform_json_to_apps
    st.info("Upload a JSON exported from **Typeform**.")
    json_lang = st.radio("Target sheet", ["English (All Anglais)", "Français (All Français)"],
                         horizontal=True, key="json_lang")
    new_json = st.file_uploader("Upload JSON", type=["json"], key="new_json")
    if new_json:
        new_apps_json = typeform_json_to_apps(new_json.read().decode("utf-8"))
        st.info(f"Found **{len(new_apps_json)}** application(s) in JSON")
        if st.button("✅ Evaluate & Add all", key="btn_json", type="primary"):
            _process_batch(
                new_apps_json,
                "All Anglais" if json_lang.startswith("English") else "All Français",
                EN_COLS if json_lang.startswith("English") else FR_COLS,
                wb, existing_names, has_fs, uploaded_xl_name,
            )

with tab_override:
    st.markdown("#### 🎛️ Manual Decision Override")
    st.caption("All startups are shown below. Change any decision using the dropdown — only rows you modify will be saved.")

    OVERRIDE_OPTIONS = ["Selected ★", "Selected", "Shortlisted ✓", "Shortlisted", "Rejected"]

    # Build full deduplicated list
    all_apps_override = []
    if has_en:
        for a in _extract_apps(wb["All Anglais"], EN_COLS):
            all_apps_override.append((a, "All Anglais", EN_COLS))
    if has_fr:
        for a in _extract_apps(wb["All Français"], FR_COLS):
            all_apps_override.append((a, "All Français", FR_COLS))

    seen_ov = set()
    unique_override = []
    for item in all_apps_override:
        key = item[0]['startup_name'].lower().strip()
        if key not in seen_ov:
            seen_ov.add(key)
            unique_override.append(item)

    # Evaluate all
    evaluated_override = [(a, evaluate(a), sheet, cm) for a, sheet, cm in unique_override]

    # ── Filters bar ────────────────────────────────────────────────────────
    fc1, fc2, fc3 = st.columns([2, 2, 3])
    with fc1:
        filter_decision = st.multiselect(
            "Filter by decision",
            options=OVERRIDE_OPTIONS,
            default=[],
            key="ov_filter_dec",
            placeholder="All decisions"
        )
    with fc2:
        filter_sector = st.multiselect(
            "Filter by sector",
            options=sorted(set(a['sector'] for a, _, _, _ in evaluated_override if a['sector'])),
            default=[],
            key="ov_filter_sec",
            placeholder="All sectors"
        )
    with fc3:
        search_text = st.text_input("🔍 Search startup name", key="ov_search", placeholder="Type to filter...")

    # Apply filters
    filtered = evaluated_override
    if filter_decision:
        filtered = [(a, r, s, cm) for a, r, s, cm in filtered if r.final_decision in filter_decision]
    if filter_sector:
        filtered = [(a, r, s, cm) for a, r, s, cm in filtered if a['sector'] in filter_sector]
    if search_text:
        filtered = [(a, r, s, cm) for a, r, s, cm in filtered
                    if search_text.lower() in a['startup_name'].lower()]

    st.caption(f"Showing **{len(filtered)}** of **{len(evaluated_override)}** startups")
    st.divider()

    # ── Table header ────────────────────────────────────────────────────────
    hcols = st.columns([3, 2, 2, 2, 3, 3, 3])
    for col, label in zip(hcols, ["**Startup**", "**Sector**", "**Age**",
                                   "**Bonus**", "**Engine Decision**", "**Reason**", "**Override To**"]):
        col.markdown(label)
    st.divider()

    # ── One row per startup ─────────────────────────────────────────────────
    pending_overrides = {}  # name → (new_decision, app, sheet, cm)

    for idx, (a, r, sheet, cm) in enumerate(filtered):
        row_cols = st.columns([3, 2, 2, 2, 3, 3, 3])
        row_cols[0].markdown(f"**{a['startup_name']}**  \n<span style='font-size:11px;color:#888'>{a.get('governorate','')}</span>", unsafe_allow_html=True)
        row_cols[1].markdown(f"<span style='font-size:12px'>{a.get('sector') or '—'}</span>", unsafe_allow_html=True)
        row_cols[2].markdown(f"<span style='font-size:12px'>{a.get('age') or '—'}</span>  \n<span style='font-size:11px;color:#aaa'>{a.get('maturity','')}</span>", unsafe_allow_html=True)

        # Engine decision badge
        eng_dec = r.final_decision
        c_eng = DECISION_COLORS.get(eng_dec, "#888")
        row_cols[4].markdown(
            f"<span style='background:{c_eng}22;color:{c_eng};font-weight:700;"
            f"padding:3px 8px;border-radius:4px;font-size:12px'>{eng_dec}</span>",
            unsafe_allow_html=True
        )

        # Reason column — rejection reason OR bonus details for accepted
        if r.rejection_reason:
            reason_html = f"<span style='color:#ff4d6d;font-size:11px'>✗ {r.rejection_reason}</span>"
        elif r.bonus_details:
            pts = " · ".join(r.bonus_details)
            reason_html = f"<span style='color:#00e5a0;font-size:11px'>✓ {pts}</span>"
        elif eng_dec in ("Shortlisted ✓", "Shortlisted"):
            filters_ok = []
            if r.filter1_result and "Pass" in str(r.filter1_result): filters_ok.append("Age ✓")
            if r.filter2_result and "Pass" in str(r.filter2_result): filters_ok.append("Legal ✓")
            if r.filter3_result and "Pass" in str(r.filter3_result): filters_ok.append("Founder ✓")
            if r.filter4_result and "Pass" in str(r.filter4_result): filters_ok.append("Maturity ✓")
            reason_html = f"<span style='color:#f5c842;font-size:11px'>{' · '.join(filters_ok) or 'Filters passed'}</span>"
        else:
            reason_html = "<span style='color:#666;font-size:11px'>—</span>"
        row_cols[5].markdown(reason_html, unsafe_allow_html=True)

        # Override dropdown — default to engine decision
        default_idx = OVERRIDE_OPTIONS.index(eng_dec) if eng_dec in OVERRIDE_OPTIONS else 0
        new_dec = row_cols[6].selectbox(
            label="override",
            options=OVERRIDE_OPTIONS,
            index=default_idx,
            key=f"ov_{idx}_{a['startup_name']}",
            label_visibility="collapsed",
        )

        # Track if changed
        if new_dec != eng_dec:
            pending_overrides[a['startup_name']] = (new_dec, a, sheet, cm)

    # ── Summary of pending changes ──────────────────────────────────────────
    st.divider()
    if pending_overrides:
        st.markdown(f"#### ✏️ {len(pending_overrides)} pending override(s):")
        for name, (dec, _, _, _) in pending_overrides.items():
            c = DECISION_COLORS.get(dec, "#888")
            st.markdown(
                f"- **{name}** → <span style='color:{c};font-weight:700'>{dec}</span>",
                unsafe_allow_html=True
            )

        if st.button("💾 Apply All Overrides", type="primary", key="btn_apply_all_overrides"):
            applied = []
            for name, (new_decision, app_ov, sheet_ov, cm_ov) in pending_overrides.items():
                ws_ov = wb[sheet_ov]
                name_col_idx = cm_ov["startup_name"]

                for row_idx in range(2, ws_ov.max_row + 1):
                    cell_val = ws_ov.cell(row=row_idx, column=name_col_idx).value
                    if cell_val and str(cell_val).strip().lower() == name.lower():
                        headers = [ws_ov.cell(row=1, column=c).value for c in range(1, ws_ov.max_column + 1)]
                        if "Manual Override" in headers:
                            ov_col_idx = headers.index("Manual Override") + 1
                        else:
                            ov_col_idx = ws_ov.max_column + 1
                            ws_ov.cell(row=1, column=ov_col_idx).value = "Manual Override"
                            ws_ov.cell(row=1, column=ov_col_idx).font = Font(bold=True)
                        ws_ov.cell(row=row_idx, column=ov_col_idx).value = new_decision
                        applied.append(name)
                        break

                # Sync Final selection sheet
                if has_fs:
                    ws_fs = wb["Final selection"]
                    fs_name_col = FS_COLS["startup_name"]
                    in_final = any(
                        ws_fs.cell(row=r2, column=fs_name_col).value and
                        str(ws_fs.cell(row=r2, column=fs_name_col).value).strip().lower() == name.lower()
                        for r2 in range(2, ws_fs.max_row + 1)
                    )
                    is_now_selected = new_decision in ("Selected ★", "Selected")

                    if is_now_selected and not in_final:
                        _append_to_final_selection(ws_fs, app_ov)
                    elif not is_now_selected and in_final:
                        for r2 in range(2, ws_fs.max_row + 1):
                            fv = ws_fs.cell(row=r2, column=fs_name_col).value
                            if fv and str(fv).strip().lower() == name.lower():
                                ws_fs.delete_rows(r2)
                                break

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            st.success(f"✅ {len(applied)} override(s) applied: " + ", ".join(applied))
            stem = uploaded_xl_name.replace(".xlsx", "")
            st.download_button(
                "⬇️ Download Updated Excel",
                data=buf.getvalue(),
                file_name=f"{stem}_Updated.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
    else:
        st.info("No changes yet — adjust any dropdown above to create an override.")

with tab_manual:
    lang = st.radio("Target sheet",
                    ["English (All Anglais)", "Français (All Français)"],
                    horizontal=True, key="manual_lang")
    is_en = lang.startswith("English")

    SECTORS = [
        "AI/Data", "FinTech", "HealthTech", "EdTech", "GreenTech",
        "Agri/FoodTech", "HR Tech", "DeepTech", "SaaS", "TravelTech",
        "PropTech", "Industry 4.0", "ICC", "Transport Tech",
        "Industries créatives & culturelles (ICC)", "Other",
    ]
    GOVERNORATES = [
        "Tunis", "Ariana", "Ben Arous", "Manouba", "Nabeul", "Bizerte",
        "Sousse", "Monastir", "Sfax", "Kairouan", "Mahdia", "Médenine",
        "Gafsa", "Jendouba", "Other",
    ]
    LEGAL_TYPES = ["SARL", "SUARL", "SA", "Association", "Auto-entrepreneur", "Other"]
    HIRING_OPTIONS = ["None", "1–2", "3–5", "6–10", "+10"]
    FOUNDER_COUNT = ["1", "2", "3 et plus"]
    GENDER_OPTIONS = ["Male", "Female"]
    AGE_FOUNDER_OPTIONS = ["Less than 25 years", "25–34 years", "35–44 years", "45+ years"]

    with st.form("add_startup_form"):
        st.markdown("#### 🏢 Startup Info")
        c1, c2, c3 = st.columns(3)
        with c1:
            name        = st.text_input("Startup name *")
            age         = st.selectbox("Age of startup *", AGE_OPTIONS)
            founded_year= st.text_input("Year founded (e.g. 2023)")
        with c2:
            governorate = st.selectbox("Governorate", GOVERNORATES)
            sector      = st.selectbox("Sector", SECTORS)
            legal_type  = st.selectbox("Legal type", LEGAL_TYPES)
        with c3:
            website     = st.text_input("Website")
            linkedin    = st.text_input("LinkedIn")
            facebook    = st.text_input("Facebook")

        st.markdown("#### 👤 Founder(s)")
        num_founders = st.selectbox("Number of founders", FOUNDER_COUNT)
        fc1, fc2, fc3 = st.columns(3)
        founders = []
        for i, col in enumerate([fc1, fc2, fc3], 1):
            with col:
                st.markdown(f"**Founder {i}**")
                fn = st.text_input(f"Full name", key=f"fn{i}")
                fp = st.text_input(f"Phone", key=f"fp{i}")
                fe = st.text_input(f"Email", key=f"fe{i}")
                fg = st.selectbox(f"Gender", GENDER_OPTIONS, key=f"fg{i}")
                ff = st.text_input(f"Function/Role", key=f"ff{i}")
                fa = st.selectbox(f"Age range", AGE_FOUNDER_OPTIONS, key=f"fa{i}")
                fl = st.text_input(f"LinkedIn", key=f"fl{i}")
                founders.append({"name": fn, "phone": fp, "email": fe,
                                  "gender": fg, "function": ff, "age": fa, "linkedin": fl})

        st.markdown("#### ✅ Eligibility")
        e1, e2, e3, e4 = st.columns(4)
        with e1: legally = st.checkbox("Legally established?")
        with e2: fulltime = st.checkbox("Full-time founder?")
        with e3: resident = st.checkbox("Founder in Tunisia?")
        with e4: mixed = st.checkbox("Gender-mixed team?")

        st.markdown("#### 👥 Team")
        t1, t2 = st.columns(2)
        with t1: employees = st.selectbox("Total employees", EMPLOYEE_RANGES)
        with t2: salaried  = st.selectbox("Salaried employees", EMPLOYEE_RANGES[:-1])
        hiring = st.selectbox("Hiring plans (next 12 months)", HIRING_OPTIONS)

        st.markdown("#### 💡 Solution")
        maturity = st.selectbox("Solution maturity *", MATURITY_OPTIONS)
        problem  = st.text_area("Problem you solve", height=80)
        solution = st.text_area("Solution you offer", height=80)
        tech     = st.text_area("Technology / Innovation", height=60)
        diff     = st.text_area("Differentiation from competition", height=60)
        pitch    = st.text_area("Elevator Pitch (max 5 sentences)", height=80)

        st.markdown("#### 📈 Traction")
        tr1, tr2, tr3, tr4 = st.columns(4)
        with tr1: clients  = st.selectbox("Clients / Users", CLIENT_RANGES)
        with tr2: revenue  = st.checkbox("Generating revenue?")
        with tr3: funded   = st.checkbox("Raised funding?")
        with tr4: label    = st.checkbox("Startup label?")
        total_income = st.text_input("Total income in TND (since launch)")
        rev_2025 = st.text_input("Revenue 2025")
        rev_2024 = st.text_input("Revenue 2024")
        rev_2023 = st.text_input("Revenue 2023")
        amount_raised = st.text_input("Total amount raised (if any)")

        submitted = st.form_submit_button("🚀 Evaluate & Add", type="primary", use_container_width=True)

    if submitted:
        if not name:
            st.error("Startup name is required.")
            st.stop()

        if name.lower().strip() in existing_names:
            st.warning(f"⚠️ **{name}** already exists in the file.")
            st.stop()

        app = {
            'startup_name':          name,
            'age':                   age,
            'sector':                sector,
            'governorate':           governorate,
            'legally_created':       legally,
            'full_time_founder':     fulltime,
            'founder_in_tunisia':    resident,
            'total_employees':       employees,
            'salaried_employees':    salaried,
            'gender_mixed':          mixed,
            'maturity':              maturity,
            'num_clients':           clients,
            'generating_revenue':    revenue,
            'raised_funding':        funded,
            'startup_label':         label,
            'participated_programs': False,
        }

        result = evaluate(app)

        # ── Show result ───────────────────────────────────────────────────────
        st.divider()
        st.markdown("### 🔍 Evaluation Result")

        col_res, col_det = st.columns([1, 2])
        with col_res:
            c = DECISION_COLORS.get(result.final_decision, "#888")
            st.markdown(
                f"<div style='background:{c}22;border-left:5px solid {c};"
                f"padding:16px;border-radius:8px;font-size:22px;font-weight:bold;color:{c}'>"
                f"{result.emoji} {result.final_decision}</div>",
                unsafe_allow_html=True
            )
            st.metric("Bonus Score", result.bonus_score)
        with col_det:
            st.markdown("**Filter results:**")
            st.markdown(f"- F1 Age: `{result.filter1_result}`")
            st.markdown(f"- F2 Legal: `{result.filter2_result}`")
            st.markdown(f"- F3 Founder: `{result.filter3_result}`")
            st.markdown(f"- F4 Maturity: `{result.filter4_result}`")
            if result.rejection_reason:
                st.error(f"Rejection reason: {result.rejection_reason}")
            if result.bonus_details:
                st.markdown("**Bonus points:** " + " · ".join(result.bonus_details))

        # ── Write to workbook ─────────────────────────────────────────────────
        target_sheet = "All Anglais" if is_en else "All Français"
        col_map      = EN_COLS if is_en else FR_COLS

        _append_to_main_sheet(wb[target_sheet], col_map, app, founded_year)

        is_selected = result.final_decision in ("Selected ★", "Selected")
        if is_selected and has_fs:
            _append_to_final_selection(wb["Final selection"], app, founded_year)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        if is_selected and has_fs:
            st.success(f"✅ **{name}** added to `{target_sheet}` **and** `Final selection`")
        else:
            st.success(f"✅ **{name}** added to `{target_sheet}`")

        stem = uploaded_xl_name.replace(".xlsx", "")
        st.download_button(
            "⬇️ Download Updated Excel",
            data=buf.getvalue(),
            file_name=f"{stem}_Updated.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )